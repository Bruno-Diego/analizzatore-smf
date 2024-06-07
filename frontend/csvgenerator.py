# csvgenerator_v0.2

import click
import csv
import logging
from openpyxl import Workbook, load_workbook

# Initialize the logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Create a file handler (optional)
file_handler = logging.FileHandler('csvgen.log')
file_handler.setLevel(logging.INFO)

# Define the log format
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Add handlers to the logger
logger.addHandler(file_handler)

csvcontent = []
final_data = []
user_option_yes = ["yes", "y", "si", "s"]
user_option_no = ["no", "n"]


def filter_columns(column, header, csvcontent):
    # filter columns 
    filtered_head = [header[col] for col in column]
    filtered_cols = []
    for row in csvcontent:
        filtered_cols.append([row[col] for col in column ])
    # final_data = [filtered_cols[i:i + len(filtered_head)] for i in range(0, len(filtered_cols), len(filtered_head))]
    filtered_cols.insert(0, filtered_head)
    # log the filtered colums 
    return filtered_cols


def save_csv(user_output_name, delimiter, quotechar, header, final_data):
    """
    Saves CSV data to an CSV file.
    """
    #    Prompt user for confirmation or new output file name
    with open(user_output_name, 'x', newline='', ) as output_file:
        writer = csv.writer(output_file, delimiter=delimiter, quotechar=quotechar)
        writer.writerows([header])
        writer.writerows(final_data) 
    click.echo(f"Data saved to {user_output_name}.")
    click.echo("CSV filtering completed successfully.")
    # Log successful completion
    logger.info(f"Data saved to {user_output_name}.")
    logger.info("CSV filtering completed successfully.")


def save_csv_to_excel(wb, header, final_data, user_output_name):
    """
    Saves CSV data to an Excel workbook.

    Args:
        wb (openpyxl.Workbook): The target workbook.
        header (list): List of column headers.
        final_data (list): List of lists representing CSV data.
        user_output_name (str): Name of the output Excel file.

    Returns:
        None
    """
    ws1 = wb['Sheet']
    header = final_data[0]
    for h in range(len(header)):
        ws1.cell(1, 1+h).value = header[h]
    final_data.pop(0)
    # print(header, final_data)
    for column in final_data:
        next_row = ws1.max_row + 1 # Find the next available row
        for i in range(len(column)):
            ws1.cell(next_row, i + 1).value = column[i]

    wb.save(user_output_name)
    click.echo(f"Data saved to {user_output_name}.")
    click.echo("Conversion completed successfully.")
    logger.info(f"Data saved to {user_output_name}.")
    logger.info("Conversion completed successfully.")


interactive_questions = {
    1: "Do you want to save the data as csv file? (yes/no): ", 
    2: "What will be the name of the file? (example.csv): ",
    3: "Thank you for testing our program",
    4: "That's not a valid response, please run the program again and provide a valid answer.",
    5: "The command is missing arguments, use the flag --help for instructions",
    6: "Do you want to save the data as Excel file? (yes/no): ",
    7: "What will be the name of the excel file? (example.xlsx): ",
    8: "The file already exists, want to override it's content? (Yes/No)"
    }
    
    
@click.group()
def cli():
    """
    This is a program that can read your text file and convert 
    to CSV or XLSX (Excel) formats.
    Use 'filtercsv' for csv and 'convert' for xlsx conversions.
    """
    pass


@cli.command()
@click.argument('src')
@click.option('--output', help="Give the name for the output file")
@click.option('--column', type=int, multiple=True, help="Give the column to be filtred")
@click.option('--delimiter', default=';', help="Specify the delimiter character")
@click.option('--quotechar', default='/', help="Specify the quote character")
@click.option('--interactive', is_flag=True, help="Get an interactive prompt")
def filtercsv(src, output, column, delimiter, quotechar, interactive):
    """
    Display file data, filter by columns, and convert to a CSV file.

    Args:\n
        src (str): Path to the input CSV file.\n
        output (str): Name for the output CSV file.\n
        column (List[int]): Optional. Indices of columns to filter.\n
        delimiter (str): Optional. Delimiter character (default: ;).\n
        quotechar (str): Optional. Quote character (default: ").\n
        interactive (bool): Optional. Interactive prompt.\n
        
    Example usage:
        $ csvgenerator_v2.py filtercsv input.csv --output filtered_output.csv --column 1 --column 3 --delimiter , --quotechar | --interactive
    """
    try:
        # Read the file
        if not src.endswith(".csv"):
            return click.echo(f"File not supported. Please check the file extention.")
        click.echo(f"You chose {src} to work with.")
        
        with open(src, newline='') as csvfile:
            content = csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar)
            header = next(content)
            csvcontent = [ row for row in content ]
            # click.echo(f"header: {header} \nContent: {csvcontent}")
        logger.info(f"Processing file: {src}")
    except FileNotFoundError:
        return click.echo(f"File not found. Please check the file path.")
    except Exception as e:
        return click.echo(f"An error occurred: {e}")
    try:
        if column:
            final_data = filter_columns(column, header, csvcontent)
        else:
            final_data = [header] + csvcontent
    except IndexError as e:
        return click.echo(f"Please check if the column index is correct, an error occurred: {e}.")
    except Exception as e:
        return click.echo(f"An error occurred: {e}")
    try:
        # If the user provided output, save the file with the choosen name
        if final_data and output:
            click.echo(f"output: {output}")
            with open(output, mode='x', newline='') as file:
                writer = csv.writer(file, delimiter=delimiter, quotechar="/")
                # writer.writerows([header])
                writer.writerows(final_data)
            # Log successful completion
            logger.info(f"Data saved to {output}.")
            logger.info("CSV filtering completed successfully.")
        else:
            # ask_user(output_file, delimiter, quotechar, header, final_data)
            # Prompt user for confirmation or new output file name
            if interactive:
                user_response = input(interactive_questions[1])
                if user_response in user_option_yes:
                    user_output_name = input(interactive_questions[2])
                    save_csv(user_output_name, delimiter, quotechar, header, final_data)
                elif user_response in user_option_no:
                    return click.echo(interactive_questions[3])
                else:
                    return click.echo(interactive_questions[4])
            else:
                return click.echo(interactive_questions[5])
    except FileExistsError as e:
        click.echo(f"There is a file with the same name on this path. An error occurred: {e}. Choose another name.")
        if interactive:
            user_output_name = input(interactive_questions[2])
            save_csv(user_output_name, delimiter, quotechar, header, final_data)
    except AssertionError:
        click.echo("Missing one or more arguments")
    except Exception as e:
        return click.echo(f"An error occurred: {e}")


@cli.command()
@click.argument('src')
@click.option('--output', help="Specify the output Excel file name")
@click.option('--column', type=int, multiple=True, help="Specify the column(s) to include in the Excel file")
@click.option('--delimiter', default=';', help="Specify the delimiter character")
@click.option('--quotechar', default='"', help="Specify the quote character")
@click.option('--interactive', is_flag=True, help="Get an interactive prompt")
def convert(src, output, column, delimiter, quotechar, interactive):
    """
    Convert data from a CSV file to an Excel (xlsx) file.

    Args:\n
        src (str): Path to the input CSV file.\n
        output (str): Optional. Name for the output Excel file.\n
        column (List[int]): Optional. Indices of columns to include.\n
        delimiter (str): Optional. Delimiter character (default: ;).\n
        quotechar (str): Optional. Quote character (default: \").\n
        interactive (bool): Optional. Interactive prompt.\n

    Example usage:
        $ csvgenerator_v2.py convert input.csv --output converted_output.xlsx --column 1 --column 3 --delimiter , --quotechar | --interactive
    """
    try:
        # Read the file
        click.echo(f"You chose {src} to work with.")
        if not src.endswith(".csv"):
                return click.echo(f"File not supported. Please check the file extention.")
        with open(src, newline='') as csvfile:
            content = csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar)
            header = next(content)
            csvcontent = [ row for row in content ]
        logger.info(f"Processing file: {src}")
    except FileNotFoundError:
        return click.echo(f"File not found. Please check the file path.")
    except Exception as e:
        return click.echo(f"An error occurred: {e}")
    try:
        if column:
            final_data = filter_columns(column, header, csvcontent)
        else:
            final_data = [header] + csvcontent
            print(final_data)
    except IndexError as e:
        return click.echo(f"Please check if the column index is correct, an error occurred: {e}.")
    except Exception as e:
        return click.echo(f"An error occurred: {e}")
    try:
        # If the user provided output, seve the file with the choosen name
        if final_data and output:
            # Load the excel workbook
            try:
                wb = load_workbook(output)
                if interactive:
                    choice = input(interactive_questions[8])
                    if choice in user_option_no:
                        user_output_name = input("What will be the new name of the file? (example.xlsx): ")
                        wb = Workbook()
                        wb.create_sheet("Sheet")
                    else:
                        return click.echo("The file already exists, please choose another name and run the program again.")
            except FileNotFoundError:
                wb = Workbook()
                wb.create_sheet("Sheet")
            user_output_name = output
            save_csv_to_excel(wb, header, final_data, user_output_name)
        else:
            if interactive:
            # Prompt user for confirmation or new output file name
                user_response = input(interactive_questions[6])
                if user_response in user_option_yes:
                    user_output_name = input(interactive_questions[7])
                    try:
                        wb = load_workbook(user_output_name)
                        choice = input(interactive_questions[8])
                        if choice in user_option_no:
                            user_output_name = input(interactive_questions[7])
                            wb = Workbook()
                            wb.create_sheet("Sheet")
                    except FileNotFoundError:
                        wb = Workbook()
                        wb.create_sheet("Sheet")
                    save_csv_to_excel(wb, header, final_data, user_output_name)    
                else:
                    return click.echo("Thank you for testing our program")
            else:
                return click.echo(interactive_questions[5])
    except FileExistsError as e:
        if interactive: 
            click.echo(f"There is a file with the same name on this path. An error occurred: {e}. Choose another name.")
            user_output_name = input(interactive_questions[7])
            try:
                wb = load_workbook(user_output_name)
            except FileNotFoundError:
                wb = Workbook()
                wb.create_sheet("Sheet")
            save_csv_to_excel(wb, header, csvcontent, user_output_name)  
        else:
            return click.echo(f"There is a file with the same name on this path. An error occurred: {e}. Choose another name.")  
    except Exception as e:
        return click.echo(f"An error occurred: {e}")

if __name__ == '__main__':
    cli()